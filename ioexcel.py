import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook


def read_excel_data(_f):
    path = _f
    xl = pd.ExcelFile(path)

    # see all sheet names
    sheet_names = xl.sheet_names  
    _df = pd.DataFrame()
    # read a specific sheet to DataFrame
    for _sn in sheet_names:
        print(_sn)
        _dfTemp = xl.parse(_sn)
        _df = pd.concat([_df, _dfTemp], axis=0)
        
    return _df

def write_to_excel_data(_df, _sname):
    #https://stackoverflow.com/questions/42370977/how-to-save-a-new-sheet-in-an-existing-excel-file-using-pandas
    path = r"bioatmosphere_data.xlsx"

    book = load_workbook(path)
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book

    #Delete sheet - _sname if it already exists
    # Get the sheets in the currnet workbook
    sheet_names = book.sheetnames
    for i in sheet_names:
        #If the sheet by name exists _sname, delete it. We are adding a new one.
        if i==_sname:
            del book[i]

    # Convert the dataframe to an XlsxWriter Excel object.
    _df.to_excel(writer, sheet_name=_sname)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()   
    
def write_to_excel_fires(_df, _sname="fires"):
    #https://stackoverflow.com/questions/42370977/how-to-save-a-new-sheet-in-an-existing-excel-file-using-pandas
    path = r"Fires_Input.xlsx"

    book = load_workbook(path)
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    
    #_sname = "fires"
     
    #Delete sheet - "fires" if it already exists
    # Get the sheets in the currnet workbook
    sheet_names = book.sheetnames
    for i in sheet_names:
        #If the sheet by name exists "fires", delete it. We are adding a new one.
        if i==_sname:
            del book[i]

    # Convert the dataframe to an XlsxWriter Excel object.
    _df.to_excel(writer, sheet_name=_sname)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()  
    

####
#    Read data from the bioatmosphere excel file
###
def  get_bioatmosphere_data():
    path = r"bioatmosphere_data.xlsx"
    xl = pd.ExcelFile(path)

    # see all sheet names
    sheet_names = xl.sheet_names  
    bioatmosphere_dict = {}

    # read a specific sheet to DataFrame
    for _sn in sheet_names:
        print(_sn)
        _df = xl.parse(_sn)
        #Create dictionary with key, value pair  as {[latitude, longitude], dataframe}
        bioatmosphere_dict.update({_sn:_df})
    return bioatmosphere_dict

####
#    Read data from the bioatmosphere excel file
###
def  get_bioatmosphere_data_for_coord(_sname):
    path = r"bioatmosphere_data.xlsx"
    xl = pd.ExcelFile(path)

    # see all sheet names
    sheet_names = xl.sheet_names  
    bioatmosphere_dict = {}

    # read a specific sheet to DataFrame
    for _sn in sheet_names:
        print(_sn)
        if(_sn == _sname):
            _df = xl.parse(_sn)
            return True, _df
    return False, None
